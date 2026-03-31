"""
松鼠AI 异议查询助手 - 数据管理和 API 后端
支持异议库的增删改查、数据导出、AI 智能匹配等功能
"""

import json
import os
from datetime import datetime
from typing import List, Dict, Optional
import re

class ObjectionAssistant:
    """异议库管理系统"""
    
    def __init__(self, data_file="objections_db.json"):
        self.data_file = data_file
        self.objections = self._load_data()
    
    def _load_data(self) -> List[Dict]:
        """加载异议数据"""
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []
    
    def _save_data(self):
        """保存异议数据"""
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(self.objections, f, ensure_ascii=False, indent=2)
    
    def add_objection(self, objection_data: Dict) -> Dict:
        """添加新异议"""
        new_id = max([o.get('id', 0) for o in self.objections] + [0]) + 1
        
        objection = {
            'id': new_id,
            'title': objection_data.get('title'),
            'category': objection_data.get('category'),
            'keywords': objection_data.get('keywords', []),
            'description': objection_data.get('description'),
            'analysis': objection_data.get('analysis'),
            'solutions': objection_data.get('solutions', []),
            'dataPoints': objection_data.get('dataPoints', []),
            'counterExamples': objection_data.get('counterExamples', []),
            'tips': objection_data.get('tips', []),
            'createdAt': datetime.now().isoformat(),
            'updatedAt': datetime.now().isoformat(),
            'usageCount': 0,
            'rating': 0
        }
        
        self.objections.append(objection)
        self._save_data()
        return objection
    
    def update_objection(self, objection_id: int, updates: Dict) -> Optional[Dict]:
        """更新异议"""
        for obj in self.objections:
            if obj['id'] == objection_id:
                obj.update(updates)
                obj['updatedAt'] = datetime.now().isoformat()
                self._save_data()
                return obj
        return None
    
    def delete_objection(self, objection_id: int) -> bool:
        """删除异议"""
        original_len = len(self.objections)
        self.objections = [o for o in self.objections if o['id'] != objection_id]
        if len(self.objections) < original_len:
            self._save_data()
            return True
        return False
    
    def search_objections(self, keyword: str) -> List[Dict]:
        """搜索异议"""
        keyword = keyword.lower()
        results = []
        
        for obj in self.objections:
            # 在标题、描述、关键词中搜索
            if (keyword in obj['title'].lower() or
                keyword in obj['description'].lower() or
                keyword in obj['category'].lower() or
                any(keyword in k.lower() for k in obj['keywords'])):
                results.append(obj)
        
        return results
    
    def get_by_category(self, category: str) -> List[Dict]:
        """按分类获取异议"""
        return [o for o in self.objections if o['category'] == category]
    
    def get_all_categories(self) -> List[str]:
        """获取所有分类"""
        return sorted(list(set(o['category'] for o in self.objections)))
    
    def smart_match(self, user_input: str) -> Dict:
        """
        AI 智能匹配 - 根据用户输入智能推荐异议
        返回最相关的异议和建议话术
        """
        # 分词和关键词提取
        keywords = self._extract_keywords(user_input)
        
        # 匹配异议
        scored_results = []
        for obj in self.objections:
            score = 0
            
            # 标题匹配
            if any(kw in obj['title'].lower() for kw in keywords):
                score += 5
            
            # 关键词匹配
            matched_keywords = [kw for kw in obj['keywords'] if any(uk in kw.lower() for uk in keywords)]
            score += len(matched_keywords) * 3
            
            # 描述匹配
            if any(kw in obj['description'].lower() for kw in keywords):
                score += 2
            
            if score > 0:
                scored_results.append((obj, score))
        
        # 排序并返回Top 3
        scored_results.sort(key=lambda x: x[1], reverse=True)
        
        return {
            'input': user_input,
            'keywords': keywords,
            'recommendations': [
                {
                    'objection': obj,
                    'score': score,
                    'matchReason': self._generate_match_reason(obj, keywords)
                }
                for obj, score in scored_results[:3]
            ]
        }
    
    def _extract_keywords(self, text: str) -> List[str]:
        """提取关键词"""
        # 简单的分词（实际应用中建议使用 jieba）
        text = text.lower()
        # 移除标点
        text = re.sub(r'[，。！？；：""''（）【】]', ' ', text)
        # 分词
        words = text.split()
        # 过滤短词和停用词
        stopwords = {'的', '是', '我', '你', '他', '和', '在', '吗', '呢', '啊', '吧', '了', '、'}
        keywords = [w for w in words if len(w) > 1 and w not in stopwords]
        return keywords
    
    def _generate_match_reason(self, obj: Dict, keywords: List[str]) -> str:
        """生成匹配理由"""
        matched = [kw for kw in obj['keywords'] if any(uk in kw.lower() for uk in keywords)]
        if matched:
            return f"匹配关键词：{', '.join(matched[:2])}"
        return "相关度较高"
    
    def record_usage(self, objection_id: int):
        """记录异议被查询的次数"""
        for obj in self.objections:
            if obj['id'] == objection_id:
                obj['usageCount'] = obj.get('usageCount', 0) + 1
                self._save_data()
                break
    
    def rate_solution(self, objection_id: int, rating: int):
        """评分解决方案"""
        for obj in self.objections:
            if obj['id'] == objection_id:
                obj['rating'] = rating
                obj['ratedAt'] = datetime.now().isoformat()
                self._save_data()
                break
    
    def get_statistics(self) -> Dict:
        """获取统计数据"""
        if not self.objections:
            return {
                'totalObjections': 0,
                'totalCategories': 0,
                'totalUsage': 0,
                'averageRating': 0,
                'categoryDistribution': {},
                'mostUsed': [],
                'topRated': []
            }
        
        total_usage = sum(o.get('usageCount', 0) for o in self.objections)
        avg_rating = sum(o.get('rating', 0) for o in self.objections) / len(self.objections) if self.objections else 0
        
        # 分类分布
        category_dist = {}
        for obj in self.objections:
            cat = obj['category']
            category_dist[cat] = category_dist.get(cat, 0) + 1
        
        # 最常用
        most_used = sorted(self.objections, key=lambda x: x.get('usageCount', 0), reverse=True)[:5]
        
        # 高评分
        top_rated = sorted(self.objections, key=lambda x: x.get('rating', 0), reverse=True)[:5]
        
        return {
            'totalObjections': len(self.objections),
            'totalCategories': len(set(o['category'] for o in self.objections)),
            'totalUsage': total_usage,
            'averageRating': round(avg_rating, 2),
            'categoryDistribution': category_dist,
            'mostUsed': [{'id': o['id'], 'title': o['title'], 'count': o.get('usageCount', 0)} for o in most_used],
            'topRated': [{'id': o['id'], 'title': o['title'], 'rating': o.get('rating', 0)} for o in top_rated]
        }
    
    def export_to_excel(self, filename="objections_export.xlsx"):
        """导出异议库到 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("需要安装 openpyxl: pip install openpyxl")
            return False
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "异议库"
        
        # 设置表头
        headers = ['ID', '异议标题', '分类', '异议描述', '分析', '回复方案数', '数据点数', '使用次数', '评分']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加数据
        for obj in self.objections:
            ws.append([
                obj['id'],
                obj['title'],
                obj['category'],
                obj['description'][:50] + "..." if len(obj['description']) > 50 else obj['description'],
                obj['analysis'][:50] + "..." if len(obj['analysis']) > 50 else obj['analysis'],
                len(obj.get('solutions', [])),
                len(obj.get('dataPoints', [])),
                obj.get('usageCount', 0),
                obj.get('rating', 0)
            ])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 35
        
        wb.save(filename)
        print(f"✓ 已导出到 {filename}")
        return True
    
    def export_to_json(self, filename="objections_export.json"):
        """导出异议库到 JSON"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.objections, f, ensure_ascii=False, indent=2)
        print(f"✓ 已导出到 {filename}")
        return True


# 使用示例
if __name__ == "__main__":
    assistant = ObjectionAssistant()
    
    # 示例：智能匹配
    print("=" * 60)
    print("智能匹配示例")
    print("=" * 60)
    
    test_input = "客户说价格太贵了，觉得竞品便宜"
    result = assistant.smart_match(test_input)
    
    print(f"用户输入：{result['input']}")
    print(f"提取关键词：{result['keywords']}")
    print(f"\n推荐异议：")
    for rec in result['recommendations']:
        obj = rec['objection']
        print(f"\n  📌 {obj['title']}")
        print(f"     分类：{obj['category']}")
        print(f"     匹配度：{rec['score']}")
        print(f"     理由：{rec['matchReason']}")
    
    # 获取统计
    print("\n" + "=" * 60)
    print("统计数据")
    print("=" * 60)
    stats = assistant.get_statistics()
    print(f"总异议数：{stats['totalObjections']}")
    print(f"异议分类：{stats['totalCategories']}")
    print(f"总查询次数：{stats['totalUsage']}")
    print(f"平均评分：{stats['averageRating']}/5")
    print(f"\n分类分布：{stats['categoryDistribution']}")
